from multiprocessing.sharedctypes import Value
from plistlib import InvalidFileException
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import tkinter as tk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
from tkinter import ttk
from datetime import date
import os
import sys

orders = False
routes = False
saveDirectory = False

check1 = False
check2 = False
check3 = False

today = date.today()

def logisticsTool():
    global routes
    global orders
    global saveDirectory

    bold = Font(bold=True)
    italics = Font(italic=True)
    alignLeft = Alignment(horizontal='left')

    #getting order data 
    #this is putting restaurant name as key and the values as [(item, quantity)]
    orderWs = orders['Sheet1']
    ordersDict = {}
    for row in orderWs.iter_rows(values_only=True):
        if row[8] != None:
            if row[8] not in ordersDict:
                ordersDict[row[8].replace(" -  ", " - ")] = [(row[10], row[14])]
            else:
                ordersDict[row[8]].append((row[10], row[14]))

    #getting route data and setting up workbook sheets
    routes = routes[['Date', 'Driver', 'Vehicle', 'Location', 'Address']]
    drivers = routes['Driver'].unique()
    result = Workbook()
    for driver in drivers:
        result.create_sheet(driver)
    del result['Sheet']

    #putting the data together
    for sheet in result:
        sheetname = sheet.title 
        sheetDf = routes.loc[(routes['Driver'] == sheetname)]
        sheet['A1']='Date'
        sheet['A2']='Driver'
        sheet['A3']='Vehicle'
        sheet['B1']=sheetDf.iat[0,0].date()
        sheet['B2']=sheet.title
        sheet['B3']=sheetDf.iat[0,2]
        sheet['B1'].font = bold
        sheet['B2'].font = bold
        sheet['B3'].font = bold
        sheet['B1'].alignment = alignLeft
        sheet['A5']='Stop #'
        sheet['B5']='Item'
        sheet['C5']='Qty'
        sheet.column_dimensions['A'].width = 7
        sheet.column_dimensions['B'].width = 75
        sheet.column_dimensions['C'].width = 3.75
        r = 6
        stopNum = 1
        for index, row in sheetDf.iterrows():
            nameAddress = row['Location'].split(',', 1)
            if len(nameAddress) < 2 and (row['Location'] != row['Address']):
                nameAddress.append(row['Address'])
            else:
                nameAddress.append("Location does not have an address")
            name = nameAddress[0]
            address = nameAddress[1]
            sheet.cell(row=r, column=1).value = stopNum
            stopNum = stopNum + 1
            sheet.cell(row=r, column=2).value = name
            sheet.cell(row=r, column=2).font = bold
            r = r+1
            sheet.cell(row=r, column=2).value = address
            sheet.cell(row=r, column=2).font = italics
            r = r+1

            if name not in ordersDict:
                sheet.cell(row=r, column=2).value = 'Location does not have orders on Quickbooks file'
                r = r+1
                continue
            for itemQtyTuple in ordersDict[name]:
                sheet.cell(row=r, column=2).value = itemQtyTuple[0]
                sheet.cell(row=r, column=3).value = itemQtyTuple[1]
                r = r+1
            r = r+1 
    result.save(f'{os.path.basename(saveAs)}.xlsx')

#graphical user interface
root = tk.Tk()
root.title('Logistics Tool')
root.resizable(False, False)
root.geometry('700x600')
os.chdir(sys._MEIPASS)
root.iconbitmap('aaaicon.ico')

#Logistics Tool and description label
frame1 = tk.Frame(root)
frame1.place(relx=0, rely=0, relheight=0.2, relwidth=1, anchor='nw')
label1a = tk.Label(frame1, text='Logistics Tool', font=('Calibri', 40), fg='#59bab4')
label1a.place(relx=0.05, rely=0, relheight=1, relwidth=0.4)

description = """
The Logistics Tool combines data from 
Quickbooks and OptimoRoute documents to 
create a single print ready .xlsx file
that has Route and Order information."""

label1b = tk.Label(frame1, text=description, font=('Calibri', 12), fg='#707070')
label1b.place(relx=0.5, rely=0, relheight=1, relwidth=0.5)

#Select Route File label and button
frame2 = tk.Frame(root)
frame2.place(relx=0, rely=0.2, relheight=0.2, relwidth=1)
label2a = tk.Label(frame2, text='Select Route File', font=('Calibri', 20), fg='#707070')
label2a.place(relx=0, rely=0, relheight=0.5, relwidth=0.3)
label2b = tk.Label(frame2, text='(.xls \'raw\' from OptimoRoute)', font=('Calibri', 12), fg='#707070')
label2b.place(relx=0.3, rely=0, relheight=0.5, relwidth=0.3)
label2c = tk.Label(frame2, font=('Calibri', 12), fg='#707070')
label2c.place(anchor='c', relx=0.5, rely=0.5)

def selectRouteFile():
    global check1
    check1 = True
    try: 
        filetypes = (('xls files (*.xls)', '*.xls'),('All files', '*.*'))
        filename = fd.askopenfilename(title='Open a file',initialdir='/',filetypes=filetypes)
        label2c.config(text=filename)
        global routes
        routes = pd.read_excel(filename)
        check1 = True
    except:
        showinfo(title='ERROR!!!', message='Invalid File')
        check1 = False

selectRoute = ttk.Button(frame2, text='Select file', command=selectRouteFile)
selectRoute.place(relx=0.7, rely=0.17)

#Select Order File label and button
frame3 = tk.Frame(root)
frame3.place(relx=0, rely=0.4, relheight=0.2, relwidth=1)
label3a = tk.Label(frame3, text='Select Orders File', font=('Calibri', 20), fg='#707070')
label3a.place(relx=0, rely=0, relheight=0.5, relwidth=0.3)
label3b = tk.Label(frame3, text='(.xlsx from Quickbooks)', font=('Calibri', 12), fg='#707070')
label3b.place(relx=0.3, rely=0, relheight=0.5, relwidth=0.3)
label3c = tk.Label(frame3, font=('Calibri', 12), fg='#707070')
label3c.place(anchor='c', relx=0.5, rely=0.5)

def selectOrdersFile():
    global check2
    check2 = True
    try: 
        filetypes = (('xlsx files (*.xlsx)', '*.xlsx'),('All files', '*.*'))
        filename = fd.askopenfilename(title='Open a file',initialdir='/',filetypes=filetypes)
        label3c.config(text=filename)
        global orders
        orders = load_workbook(filename)
    except:
        showinfo(title='ERROR!!!', message='Invalid File')
        check2 = False

selectOrders = ttk.Button(frame3, text='Select file', command=selectOrdersFile)
selectOrders.place(relx=0.7, rely=0.17)

#Select Save Location label and button
frame4 = tk.Frame(root)
frame4.place(relx=0, rely=0.6, relheight=0.2, relwidth=1)
label4a = tk.Label(frame4, text='Save As', font=('Calibri', 20), fg='#707070')
label4a.place(relx=0, rely=0, relheight=0.5, relwidth=0.35)
label4c = tk.Label(frame4, font=('Calibri', 12), fg='#707070')
label4c.place(anchor='c', relx=0.5, rely=0.5)

def save():
    global saveAs
    global check3
    check3 = True
    files = [('xlsx files (*.xlsx)', '*.xlsx'), ('All Files', '*.*')]
    saveAs = fd.asksaveasfilename(initialfile = f'LogisticsTool {today}.xlsx', filetypes = files, defaultextension = files)
    label4c.config(text=saveAs)
    try: 
        os.chdir(os.path.dirname(saveAs))
    except OSError:
        showinfo(title='ERROR!!!', message='Please select save location')
        check3 = False
    
selectSave = ttk.Button(frame4, text='Save as', command=save)
selectSave.place(relx=0.70, rely=0.17)


#Using the tool!
frame5 = tk.Frame(root)
frame5.place(relx=0, rely=0.8, relheight=0.2, relwidth=1)
label4a = tk.Label(frame5, text='', font=('Calibri', 20), fg='#707070')
label4a.place(relx=0, rely=0, relheight=0.5, relwidth=0.35)

def runTool():
    if check1 and check2 and check3:
        try:
            logisticsTool()
        except PermissionError:
            showinfo(title='ERROR!!!', message=f'Cannot write to \n \'LogisticsTool {today}.xlsx\' \n because file with \n same name is open')
        else: 
            showinfo(title='Awesome!!!', message=f'File successfully created at: \n {saveAs}')
    else:
        showinfo(title='ERROR!!!', message='Please select valid files and/or directory')

createLogisticsFile = tk.Button(
    frame5, text='Create Logistics File!', font=('Calibri', 20), 
    fg='#707070', bg='#89f067', activebackground='#80e060', 
    command=runTool)

createLogisticsFile.place(anchor='c', relx=0.5, rely=0.5)

root.mainloop()

